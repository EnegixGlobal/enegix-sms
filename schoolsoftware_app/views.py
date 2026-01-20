from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from schoolsoftware_app.models import *
from django.contrib import messages
from django.db.models import Q, Count
from django.http import HttpResponse, HttpResponseBadRequest
from openpyxl import Workbook
from django.utils import timezone
from reportlab.pdfgen import canvas
from django.urls import reverse
from django.utils import timezone
from reportlab.lib.units import mm
from django.utils.timezone import now
from reportlab.lib.pagesizes import A4
from django.contrib.auth import logout
import datetime,os,qrcode
from django.db.models import Sum
from decimal import Decimal
import barcode
from datetime import timedelta
from datetime import date
from barcode.writer import ImageWriter
from reportlab.lib.utils import ImageReader
# from django.contrib.auth.decorators import login_required, user_passes_test
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape
from io import BytesIO
from qrcode.constants import ERROR_CORRECT_H
from django.conf import settings
# Create your views here.

# ----------------- Login ------------------------
def login_view(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = Department_loginDB.objects.filter(email=email, password=password).first()

        if user:
            request.session['user_email'] = user.email
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid email or password')
            return redirect('login')

    return render(request, 'login.html')


# ----------------- Logout ------------------------
# def logout_view(request):
#     request.session.flush()
#     return redirect('login')

def logout_view(request):
    request.session.clear()   
    request.session.flush()   
    return redirect('login')


# ---------------------------- Dashboard --------------------------------------
def dashboard_view(request):
    
    today = timezone.localdate()

    context = {
        'total_teachers': Teacher.objects.count(),
        'total_students': Student.objects.filter(status='Active').count(),
        'today_date': today,
    }
    return render(request, 'dashboard.html', context)


# All Teachers
def all_teachers(request):
    query = request.GET.get('q', '')
    teachers = Teacher.objects.filter(
        Q(name__icontains=query) |
        Q(employee_id__icontains=query) |
        Q(qualification__icontains=query)
    )
    context = {'teachers': teachers, 'query': query}
    return render(request, 'teachers/all_teachers.html', context)

# Add Teacher
def add_teacher(request):
    if request.method == 'POST':

        # Get uploaded photo
        photo = request.FILES.get('photo')

        # Create object without saving
        teacher = Teacher(
            name=request.POST.get('name'),
            dob=request.POST.get('dob'),
            gender=request.POST.get('gender'),
            phone=request.POST.get('phone'),
            email=request.POST.get('email'),
            emergency_contact=request.POST.get('emergency_contact'),
            qualification=request.POST.get('qualification'),
            experience=request.POST.get('experience'),
            employee_id=request.POST.get('employee_id'),
            joining_date=request.POST.get('joining_date'),
            salary=request.POST.get('salary'),
        )

        if photo:
            teacher.photo = photo
    
        teacher.save()

        return redirect('all_teachers')

    return render(request, 'teachers/add_teacher.html')

# Edit Teacher
def edit_teacher(request, id):
    teacher = get_object_or_404(Teacher, id=id)

    if request.method == 'POST':
        teacher.name = request.POST['name']
        teacher.dob = request.POST['dob']
        teacher.gender = request.POST['gender']
        teacher.phone = request.POST['phone']
        teacher.email = request.POST['email']
        teacher.emergency_contact = request.POST['emergency_contact']
        teacher.qualification = request.POST['qualification']
        teacher.experience = request.POST['experience']
        teacher.employee_id = request.POST['employee_id']
        teacher.joining_date = request.POST['joining_date']
        teacher.salary = request.POST['salary']
        teacher.status = request.POST.get('status', 'Active')

        # Update Photo
        if 'photo' in request.FILES:
            teacher.photo = request.FILES['photo']

        teacher.save()
        return redirect('all_teachers')

    return render(request, 'teachers/edit_teacher.html', {'teacher': teacher})


# Delete Teacher
def delete_teacher(request, id):
    teacher = get_object_or_404(Teacher, id=id)
    teacher.delete()
    return redirect('all_teachers')


# Teacher Attendance
# def teacher_attendance(request):
#     teachers = Teacher.objects.all()
#     today = datetime.date.today()

#     if request.method == 'POST':
#         date_str = request.POST.get('date', '').strip()

#         if not date_str:
#             # If date not selected, use today's date OR show error
#             date = today  
#         else:
#             date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

#         for teacher in teachers:
#             status = request.POST.get(f'status_{teacher.id}')
#             TeacherAttendance.objects.update_or_create(
#                 teacher=teacher, 
#                 date=date, 
#                 defaults={'status': status}
#             )

#         return redirect('teacher_attendance')

#     return render(request, 'teachers/teacher_attendance.html', {
#         'teachers': teachers,
#         'today': today
#     })

# def teacher_attendance(request):
#     teachers = Teacher.objects.all()
#     today = datetime.date.today()

#     if request.method == 'POST':
#         date_str = request.POST.get('date', '').strip()
#         if not date_str:
#             date = today
#         else:
#             date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

#         for teacher in teachers:
#             status = request.POST.get(f'status_{teacher.id}', 'Absent')
#             TeacherAttendance.objects.update_or_create(
#                 teacher=teacher,
#                 date=date,
#                 defaults={'status': status}
#             )

#         return redirect('teacher_attendance')

#     # Get selected date attendance if exists
#     selected_date = request.GET.get('date', today.strftime('%Y-%m-%d'))
#     attendance_records = TeacherAttendance.objects.filter(date=selected_date)
#     attendance_dict = {att.teacher.id: att.status for att in attendance_records}

#     return render(request, 'teachers/teacher_attendance.html', {
#         'teachers': teachers,
#         'today': today,
#         'attendance_dict': attendance_dict,
#         'selected_date': selected_date
#     })


# def teacher_attendance(request):
#     teachers = Teacher.objects.all()
#     today = datetime.date.today()

#     if request.method == 'POST':
#         date_str = request.POST.get('date')
#         date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

#         for teacher in teachers:
#             status = request.POST.get(f'status_{teacher.id}', 'Absent')
#             TeacherAttendance.objects.update_or_create(
#                 teacher=teacher,
#                 date=date,
#                 defaults={'status': status}
#             )

#         # ✅ IMPORTANT FIX
#         return redirect(f"{reverse('teacher_attendance')}?date={date}")

#     # -------- GET REQUEST ----------
#     selected_date = request.GET.get('date', today.strftime('%Y-%m-%d'))
#     selected_date_obj = datetime.datetime.strptime(selected_date, '%Y-%m-%d').date()

#     attendance_records = TeacherAttendance.objects.filter(date=selected_date_obj)
#     attendance_dict = {att.teacher.id: att.status for att in attendance_records}

#     return render(request, 'teachers/teacher_attendance.html', {
#         'teachers': teachers,
#         'attendance_dict': attendance_dict,
#         'selected_date': selected_date,
#         'today': today,
#     })

import calendar

def teacher_attendance(request):
    teachers = Teacher.objects.filter(status='Active')

    month = int(request.GET.get('month', datetime.date.today().month))
    year = int(request.GET.get('year', datetime.date.today().year))

    days_in_month = calendar.monthrange(year, month)[1]
    days = list(range(1, days_in_month + 1))

    attendance_qs = TeacherAttendance.objects.filter(
        date__year=year,
        date__month=month
    )

    attendance_lookup = {}
    for att in attendance_qs:
        attendance_lookup[(att.teacher_id, att.date.day)] = att.status

    attendance_rows = []

    for teacher in teachers:
        row = {
            'teacher': teacher,
            'days': []
        }

        for d in days:
            date_obj = datetime.date(year, month, d)

            # Sunday auto holiday
            if date_obj.weekday() == 6:
                status = 'Holiday'
            else:
                status = attendance_lookup.get((teacher.id, d), '-')

            row['days'].append({
                'date': date_obj.strftime('%Y-%m-%d'),
                'status': status
            })

        attendance_rows.append(row)

    return render(request, 'teachers/teacher_attendance.html', {
        'attendance_rows': attendance_rows,
        'days': days,
        'month': month,
        'year': year,
    })


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def update_teacher_attendance(request):
    if request.method == "POST":
        teacher_id = request.POST.get('teacher_id')
        date_str = request.POST.get('date')
        status = request.POST.get('status')

        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()

        TeacherAttendance.objects.update_or_create(
            teacher_id=teacher_id,
            date=date_obj,
            defaults={'status': status}
        )

        return JsonResponse({'success': True})

    return JsonResponse({'success': False})


# Teacher Payroll
def teacher_payroll(request):
    payrolls = TeacherPayroll.objects.all()
    return render(request, 'teachers/teacher_payroll.html', {'payrolls': payrolls})



# def export_attendance_excel(request):
#     date_str = request.GET.get('date')

#     if not date_str:
#         return HttpResponse("Please select a date first.", status=400)

#     selected_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()

#     # Fetch attendance for the selected date
#     attendance_records = TeacherAttendance.objects.filter(date=selected_date)

#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = "Teacher Attendance"

#     # Excel Header
#     sheet.append(["Name", "Employee ID", "Status", "Date"])

#     for record in attendance_records:
#         sheet.append([
#             record.teacher.name,
#             record.teacher.employee_id,
#             record.status,
#             record.date.strftime("%Y-%m-%d")
#         ])

#     # Prepare response
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = f'attachment; filename="attendance_{selected_date}.xlsx"'

#     workbook.save(response)
#     return response

def export_teacher_attendance_excel(request):
    date_str = request.GET.get('date')

    # 🚨 DATE MUST EXIST
    if not date_str:
        return HttpResponse("Date parameter missing", status=400)

    try:
        selected_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # 🔥 EXACT DATA FETCH
    attendance_records = TeacherAttendance.objects.filter(date=selected_date)

    # 🚨 SAFETY CHECK
    if not attendance_records.exists():
        return HttpResponse(
            f"No attendance found for {selected_date}",
            status=404
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Teacher Attendance"

    # ✅ HEADER
    sheet.append([
        "Teacher Name",
        "Employee ID",
        "Status",
        "Date"
    ])

    # ✅ DATA
    for record in attendance_records:
        sheet.append([
            record.teacher.name,
            record.teacher.employee_id,
            record.status,
            record.date.strftime("%Y-%m-%d")
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="teacher_attendance_{selected_date}.xlsx"'
    )

    workbook.save(response)
    return response


from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.http import HttpResponse
import datetime


def export_teacher_attendance_pdf(request):
    date_str = request.GET.get('date')

    if not date_str:
        return HttpResponse("Date parameter missing", status=400)

    try:
        selected_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    attendance_records = TeacherAttendance.objects.select_related('teacher').filter(
        date=selected_date
    )

    if not attendance_records.exists():
        return HttpResponse(
            f"No teacher attendance found for {selected_date}",
            status=404
        )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="teacher_attendance_{selected_date}.pdf"'
    )

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # ✅ TITLE
    title = Paragraph(
        f"<b>Teacher Attendance Report</b><br/>Date: {selected_date}",
        styles['Title']
    )
    elements.append(title)

    elements.append(Paragraph("<br/>", styles['Normal']))

    # ✅ TABLE DATA
    data = [
        ["Teacher Name", "Employee ID", "Status", "Date"]
    ]

    for record in attendance_records:
        data.append([
            record.teacher.name,
            record.teacher.employee_id,
            record.status,
            record.date.strftime("%Y-%m-%d")
        ])

    table = Table(data, colWidths=[150, 120, 80, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),

        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


 


def generate_payroll(request):
    if request.method == "POST":
        
        month = request.POST.get("month")  
        month_date = datetime.strptime(month, "%Y-%m").date()

        teachers = Teacher.objects.all()

        for teacher in teachers:
            # Total present days count
            present_days = TeacherAttendance.objects.filter(
                teacher=teacher,
                date__year=month_date.year,
                date__month=month_date.month,
                status="Present"
            ).count()

            daily_salary = teacher.salary / 30
            calculated_salary = (present_days * daily_salary)

            payroll = TeacherPayroll.objects.update_or_create(
                teacher=teacher,
                month=month_date,
                defaults={
                    "basic_salary": teacher.salary,
                    "allowances": 0,
                    "deductions": 0,
                    "total_salary": calculated_salary,
                }
            )

        messages.success(request, "Payroll generated successfully!")
        return redirect("teacher_payroll")

    return render(request, "teachers/generate_payroll.html")

def update_payroll_status(request, payroll_id):
    payroll = TeacherPayroll.objects.get(id=payroll_id)
    payroll.status = "Paid" if payroll.status == "Pending" else "Pending"
    payroll.save()
    return redirect("teacher_payroll")

# -----------------------
# ALL STUDENTS
# -----------------------
def all_students(request):
    q = request.GET.get('q','').strip()
    cls = request.GET.get('class','')
    section = request.GET.get('section','')
    status = request.GET.get('status','')

    students = Student.objects.all().select_related('school_class','section','parent')

    if q:
        students = students.filter(Q(name__icontains=q) | Q(roll_no__icontains=q))
    if cls:
        students = students.filter(school_class_id=cls)
    if section:
        students = students.filter(section_id=section)
    if status:
        students = students.filter(status=status)

    classes = SchoolClass.objects.all()
    sections = Section.objects.all()

    # Bulk actions (promote/delete)
    if request.method == 'POST':
        action = request.POST.get('action')
        selected = request.POST.getlist('selected')
        if action == 'delete' and selected:
            Student.objects.filter(id__in=selected).delete()
            messages.success(request, f"Deleted {len(selected)} students.")
            return redirect('all_students')
        if action == 'promote' and selected:
            # naive promote: increase class to next class if exists
            next_class_id = request.POST.get('next_class')
            next_section_id = request.POST.get('next_section')
            if next_class_id:
                Student.objects.filter(id__in=selected).update(school_class_id=next_class_id, section_id=next_section_id)
                messages.success(request, f"Promoted {len(selected)} students.")
                return redirect('all_students')

    return render(request, 'students/all_students.html',{
        'students': students,
        'classes': classes,
        'sections': sections,
        'q': q, 'selected_class': cls, 'selected_section': section, 'selected_status': status
    })

# -----------------------
# ADD STUDENT
# -----------------------
def add_student(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()
    parents = Parent.objects.all()

    if request.method == 'POST':
        # Parent handling: choose existing or create new
        parent_id = request.POST.get('parent')
        if parent_id:
            parent = Parent.objects.get(id=parent_id)
        else:
            parent = Parent.objects.create(
                father_name=request.POST.get('father_name',''),
                mother_name=request.POST.get('mother_name',''),
                phone=request.POST.get('parent_phone',''),
                email=request.POST.get('parent_email',''),
                address=request.POST.get('parent_address','')
            )

        student = Student(
            name=request.POST.get('name'),
            dob=request.POST.get('dob'),
            gender=request.POST.get('gender'),
            parent=parent,
            school_class_id=request.POST.get('school_class') or None,
            section_id=request.POST.get('section') or None,
            admission_date=request.POST.get('admission_date') or datetime.date.today(),
            admission_fee=request.POST.get('admission_fee') or 0,
            status=request.POST.get('status','Active'),
        )
        # roll auto/manual
        roll_no = request.POST.get('roll_no')
        if roll_no:
            student.roll_no = roll_no
        else:
            # auto generate: CLASS-SECTION-<count+1>
            cls = student.school_class.name if student.school_class else 'C'
            sec = student.section.name if student.section else 'S'
            count = Student.objects.filter(school_class=student.school_class, section=student.section).count() + 1
            student.roll_no = f"{cls[:2].upper()}{sec[:1].upper()}{count:03d}"
        # photo
        if request.FILES.get('photo'):
            student.photo = request.FILES['photo']

        student.save()
        messages.success(request, "Student admitted successfully.")
        return redirect('all_students')

    return render(request,'students/add_student.html', {
        'classes': classes, 'sections': sections, 'parents': parents
    })

# -----------------------
# EDIT & DELETE STUDENT
# -----------------------
# def edit_student(request, id):
#     student = get_object_or_404(Student, id=id)
#     classes = SchoolClass.objects.all()
#     sections = Section.objects.filter(school_class=student.school_class) if student.school_class else Section.objects.all()
#     parents = Parent.objects.all()

#     if request.method == 'POST':
#         student.name = request.POST.get('name')
#         student.dob = request.POST.get('dob')
#         student.gender = request.POST.get('gender')
#         parent_id = request.POST.get('parent')
#         if parent_id:
#             student.parent = Parent.objects.get(id=parent_id)
#         student.school_class_id = request.POST.get('school_class') or None
#         student.section_id = request.POST.get('section') or None
#         student.roll_no = request.POST.get('roll_no')
#         student.admission_date = request.POST.get('admission_date') or student.admission_date
#         student.admission_fee = request.POST.get('admission_fee') or student.admission_fee
#         student.status = request.POST.get('status','Active')
#         if request.FILES.get('photo'):
#             student.photo = request.FILES['photo']
#         student.save()
#         messages.success(request, "Student updated.")
#         return redirect('all_students')

#     return render(request,'students/edit_student.html', {'student':student,'classes':classes,'sections':sections,'parents':parents})

def edit_student(request, id):
    student = get_object_or_404(Student, id=id)

    classes = SchoolClass.objects.all()
    parents = Parent.objects.all()

    # ✅ IMPORTANT FIX
    if student.school_class:
        sections = Section.objects.filter(school_class=student.school_class)
    else:
        sections = Section.objects.none()

    if request.method == 'POST':
        student.name = request.POST.get('name')
        student.dob = request.POST.get('dob')
        student.gender = request.POST.get('gender')

        parent_id = request.POST.get('parent')
        if parent_id:
            student.parent_id = parent_id

        student.school_class_id = request.POST.get('school_class') or None
        student.section_id = request.POST.get('section') or None
        student.roll_no = request.POST.get('roll_no')
        student.admission_date = request.POST.get('admission_date') or student.admission_date
        student.admission_fee = request.POST.get('admission_fee') or student.admission_fee
        student.status = request.POST.get('status', 'Active')

        if request.FILES.get('photo'):
            student.photo = request.FILES['photo']

        student.save()
        messages.success(request, "Student updated successfully.")
        return redirect('all_students')

    return render(request, 'students/edit_student.html', {
        'student': student,
        'classes': classes,
        'sections': sections,
        'parents': parents
    })

def delete_student(request, id):
    s = get_object_or_404(Student, id=id)
    s.delete()
    messages.success(request,"Student deleted.")
    return redirect('all_students')

# -----------------------
# STUDENT PROMOTION
# -----------------------
def student_promotion(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()
    return render(request,'students/student_promotion.html', {'classes':classes,'sections':sections})

def apply_promotion(request):
    if request.method != 'POST':
        return redirect('student_promotion')
    current_class = request.POST.get('current_class')
    current_section = request.POST.get('current_section')
    next_class = request.POST.get('next_class')
    next_section = request.POST.get('next_section')
    selected_ids = request.POST.getlist('selected')  # list of student ids
    if selected_ids and next_class:
        Student.objects.filter(id__in=selected_ids).update(school_class_id=next_class, section_id=next_section)
        messages.success(request, f"Promoted {len(selected_ids)} students.")
    else:
        messages.error(request, "Select students and next class.")
    return redirect('all_students')

# -----------------------
# STUDENT ATTENDANCE
# -----------------------
def student_attendance(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()
    selected_date_str = request.GET.get('date') or datetime.date.today().strftime('%Y-%m-%d')
    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')

    # parse date
    try:
        selected_date = datetime.datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except:
        selected_date = datetime.date.today()

    # students filtered by class/section
    students = Student.objects.filter(status='Active')
    if selected_class:
        students = students.filter(school_class_id=selected_class)
    if selected_section:
        students = students.filter(section_id=selected_section)

    # prepare existing attendance dict {student_id: status}
    att_records = StudentAttendance.objects.filter(date=selected_date, student__in=students)
    att_map = {a.student_id: a.status for a in att_records}

    if request.method == 'POST':
        # selected date from form
        date_str = request.POST.get('date')
        try:
            date_to_save = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except:
            date_to_save = datetime.date.today()

        # mark all present?
        mark_all = request.POST.get('mark_all_present')
        for s in students:
            status = request.POST.get(f'status_{s.id}', 'Absent')
            if mark_all:
                status = 'Present'
            StudentAttendance.objects.update_or_create(
                student=s, date=date_to_save,
                defaults={'status': status}
            )
        messages.success(request, "Attendance saved.")
        return redirect(f"{reverse('student_attendance')}?date={date_to_save}&class={selected_class or ''}&section={selected_section or ''}")

    return render(request,'students/student_attendance.html',{
        'students': students,
        'classes': classes,
        'sections': sections,
        'selected_date': selected_date.strftime('%Y-%m-%d'),
        'selected_class': selected_class,
        'selected_section': selected_section,
        'att_map': att_map
    })

# export attendance excel
def export_attendance_excel(request):
    date_str = request.GET.get('date')
    if not date_str:
        return HttpResponseBadRequest("Select date to export")
    try:
        sel_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return HttpResponseBadRequest("Invalid date")

    records = StudentAttendance.objects.filter(date=sel_date).select_related('student')
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance_{sel_date}"
    ws.append(['Name','Roll No','Class','Section','Status','Date'])
    for r in records:
        ws.append([
            r.student.name, r.student.roll_no,
            r.student.school_class.name if r.student.school_class else '',
            r.student.section.name if r.student.section else '',
            r.status, r.date.strftime('%Y-%m-%d')
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_{sel_date}.xlsx'
    wb.save(response)
    return response

# -----------------------
# STUDENT ID CARDS
# -----------------------

# def student_id_cards(request):
#     classes = SchoolClass.objects.all()
#     sections = Section.objects.all()
#     students = Student.objects.select_related('school_class', 'section').all()

#     selected_class = request.GET.get('class')
#     selected_section = request.GET.get('section')

#     if selected_class:
#         students = students.filter(school_class_id=selected_class)
#         sections = sections.filter(school_class_id=selected_class)

#     if selected_section:
#         students = students.filter(section_id=selected_section)

#     context = {
#         'classes': classes,
#         'sections': sections,
#         'students': students,
#         'selected_class': selected_class,
#         'selected_section': selected_section,
#     }
#     return render(request, 'students/student_id_cards.html', context)

from django.core.files import File
def student_id_cards(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()
    students = Student.objects.select_related('school_class', 'section')

    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')

    if selected_class:
        students = students.filter(school_class_id=selected_class)
        sections = sections.filter(school_class_id=selected_class)

    if selected_section:
        students = students.filter(section_id=selected_section)

    for s in students:
        if not s.barcode:
            # 🔥 FULL DETAIL PAGE URL
            detail_url = request.build_absolute_uri(
                reverse('student_detail', args=[s.id])
            )

            qr = qrcode.QRCode(
                version=3,
                error_correction=qrcode.constants.ERROR_CORRECT_H,
                box_size=12,
                border=6
            )
            qr.add_data(detail_url)
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white")

            buffer = BytesIO()
            img.save(buffer, format='PNG')
            s.barcode.save(f"student_{s.id}.png", File(buffer), save=True)

    return render(request, 'students/student_id_cards.html', {
        'classes': classes,
        'sections': sections,
        'students': students,
        'selected_class': selected_class,
        'selected_section': selected_section,
    })

def student_detail(request, id):
    student = Student.objects.select_related(
        'school_class', 'section', 'parent'
    ).get(id=id)

    return render(request, 'students/student_detail.html', {
        'student': student
    })


# def download_id_cards_pdf(request):
    stud_ids = request.GET.get('students')
    cls = request.GET.get('class')
    sec = request.GET.get('section')

    qs = Student.objects.select_related('school_class', 'section').all()

    if stud_ids:
        ids = [int(x) for x in stud_ids.split(',') if x.strip().isdigit()]
        qs = qs.filter(id__in=ids)
    else:
        if cls:
            qs = qs.filter(school_class_id=cls)
        if sec:
            qs = qs.filter(section_id=sec)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="id_cards.pdf"'
    c = canvas.Canvas(response)

    y = 800
    for s in qs:
        c.rect(40, y - 140, 520, 120)
        c.setFont("Helvetica-Bold", 14)
        c.drawString(55, y - 30, "SCHOOL NAME")

        c.setFont("Helvetica", 11)
        c.drawString(55, y - 55, f"Name: {s.name}")
        c.drawString(300, y - 55, f"Roll: {s.roll_no}")
        c.drawString(
            55, y - 75,
            f"Class: {s.school_class.name if s.school_class else ''} - "
            f"{s.section.name if s.section else ''}"
        )

        if s.photo:
            try:
                c.drawImage(s.photo.path, 450, y - 120, width=70, height=90)
            except:
                pass

        y -= 180
        if y < 150:
            c.showPage()
            y = 800

    c.save()
    return response

# def download_id_cards_pdf(request):
    stud_ids = request.GET.get('students')
    cls = request.GET.get('class')
    sec = request.GET.get('section')

    qs = Student.objects.select_related('school_class', 'section').all()

    # STUDENT IDS FILTER
    if stud_ids:
        ids = [int(x) for x in stud_ids.split(',') if x.isdigit()]
        qs = qs.filter(id__in=ids)

    # CLASS FILTER (FIXED)
    if cls and cls != 'None' and cls.isdigit():
        qs = qs.filter(school_class_id=int(cls))

    # SECTION FILTER (FIXED)
    if sec and sec != 'None' and sec.isdigit():
        qs = qs.filter(section_id=int(sec))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="id_cards.pdf"'

    c = canvas.Canvas(response, pagesize=A4)

    bg_path = os.path.join(
        os.path.dirname(__file__),
        '../static/img/schoolidcard.jpeg'
    )

    x = 20 * mm
    y = 260 * mm
    card_width = 85 * mm
    card_height = 54 * mm

    for s in qs:
        # BACKGROUND
        if os.path.exists(bg_path):
            c.drawImage(bg_path, x, y, width=card_width, height=card_height)

        # TEXT
        c.setFont("Helvetica-Bold", 10)
        c.drawString(x + 25*mm, y + 35*mm, s.name)

        c.setFont("Helvetica", 9)
        c.drawString(
            x + 25*mm,
            y + 28*mm,
            f"Class: {s.school_class.name} - {s.section.name}"
        )
        c.drawString(x + 25*mm, y + 22*mm, f"Roll: {s.roll_no}")

        # PHOTO
        if s.photo and os.path.exists(s.photo.path):
            c.drawImage(
                s.photo.path,
                x + 5*mm,
                y + 20*mm,
                width=18*mm,
                height=22*mm
            )

        # NEXT CARD
        x += card_width + 10*mm
        if x + card_width > 200*mm:
            x = 20*mm
            y -= card_height + 15*mm

        if y < 40*mm:
            c.showPage()
            x = 20*mm
            y = 260*mm

    c.save()
    return response

PVC_SIZE = landscape((86 * mm, 54 * mm))

def download_id_cards_pdf(request):
    cls = request.GET.get('class')
    sec = request.GET.get('section')

    students = Student.objects.select_related(
        'school_class', 'section', 'parent'
    )

    if cls and cls.isdigit():
        students = students.filter(school_class_id=int(cls))
    if sec and sec.isdigit():
        students = students.filter(section_id=int(sec))

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Student_ID_Cards_PVC.pdf"'

    c = canvas.Canvas(response, pagesize=PVC_SIZE)

    bg_path = os.path.join(
        settings.BASE_DIR,
        'static/img/schoolidcard.jpeg'
    )

    for s in students:

        # ===== BACKGROUND (SAME AS HTML CARD) =====
        if os.path.exists(bg_path):
            c.drawImage(bg_path, 0, 0, 86*mm, 54*mm)

        # ===== PHOTO (SAME POSITION) =====
        if s.photo and os.path.exists(s.photo.path):
            c.drawImage(
                s.photo.path,
                5*mm, 23*mm,
                18*mm, 22*mm,
                mask='auto'
            )

        # ===== SCHOOL NAME =====
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawCentredString(43*mm, 49*mm, "ABC PUBLIC SCHOOL")

        # ===== STUDENT DETAILS (MATCH HTML) =====
        c.setFont("Helvetica-Bold", 8.5)
        c.drawString(26*mm, 41*mm, f"Name: {s.name}")

        c.setFont("Helvetica", 7.5)
        c.drawString(26*mm, 37*mm, f"Roll No: {s.roll_no}")
        c.drawString(
            26*mm, 33*mm,
            f"Class: {s.school_class.name} - {s.section.name}"
        )
        c.drawString(
            26*mm, 29*mm,
            f"DOB: {s.dob.strftime('%d-%m-%Y')}"
        )
        c.drawString(
            26*mm, 25*mm,
            f"Gender: {s.gender}"
        )

        if s.parent:
            c.drawString(
                5*mm, 20*mm,
                f"Parent: {s.parent.father_name}"
            )

        c.drawString(
            5*mm, 16*mm,
            f"Admission: {s.admission_date.strftime('%d-%m-%Y')}"
        )

        # ===== QR CODE (SAME LOOK + SCANNABLE) =====
        qr_data = request.build_absolute_uri(
            f"/student/{s.id}/"
        )

        qr = qrcode.QRCode(
            version=2,
            error_correction=ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(qr_data)
        qr.make(fit=True)

        qr_img = qr.make_image(fill_color="black", back_color="white")

        qr_buffer = BytesIO()
        qr_img.save(qr_buffer, format='PNG')
        qr_buffer.seek(0)

        # WHITE BASE (PVC SAFE)
        c.setFillColor(colors.white)
        c.rect(63*mm, 21*mm, 20*mm, 20*mm, fill=1, stroke=0)

        c.drawImage(
            ImageReader(qr_buffer),
            64*mm, 22*mm,
            18*mm, 18*mm
        )

        c.showPage()

    c.save()
    return response


def student_qr_profile(request, id):
    student = get_object_or_404(Student, id=id)

    return render(
        request,
        'students/student_qr_profile.html',
        {'student': student}
    )

# -----------------------
# STUDENT EXPORTS (Excel + PDF list)
# -----------------------
# def export_students_excel(request):
#     qs = Student.objects.all().select_related('school_class','section')
#     wb = Workbook()
#     ws = wb.active
#     ws.append(['Name','Roll No','Class','Section','DOB','Contact','Status'])
#     for s in qs:
#         ws.append([s.name, s.roll_no, s.school_class.name if s.school_class else '', s.section.name if s.section else '', s.dob.strftime('%Y-%m-%d'), s.parent.phone if s.parent else '', s.status])
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=students.xlsx'
#     wb.save(response)
#     return response

def export_students_excel(request):
    qs = Student.objects.select_related('school_class','section','parent').all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append(['Name','Roll No','Class','Section','DOB','Parent Phone','Status'])

    for s in qs:
        ws.append([
            s.name,
            s.roll_no,
            s.school_class.name if s.school_class else '',
            s.section.name if s.section else '',
            s.dob.strftime('%Y-%m-%d') if s.dob else '',
            s.parent.phone if s.parent else '',
            s.status
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=students.xlsx'
    wb.save(response)
    return response

# def export_students_pdf(request):
    qs = Student.objects.all().select_related('school_class','section')
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=students_list.pdf'
    c = canvas.Canvas(response)
    y = 800
    c.setFont("Helvetica-Bold",14); c.drawString(50,y,"Students List"); y -= 30
    c.setFont("Helvetica",10)
    for s in qs:
        c.drawString(50,y, f"{s.roll_no} - {s.name} - {s.school_class.name if s.school_class else ''} - {s.section.name if s.section else ''}")
        y -= 18
        if y < 80:
            c.showPage(); y = 800
    c.showPage(); c.save()
    return response


def export_students_pdf(request):
    qs = Student.objects.select_related('school_class','section').all()
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=students_list.pdf'
    c = canvas.Canvas(response)

    y = 800
    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "Students List")
    y -= 30

    c.setFont("Helvetica", 10)
    for s in qs:
        c.drawString(
            50, y,
            f"{s.roll_no} | {s.name} | "
            f"{s.school_class.name if s.school_class else ''} - "
            f"{s.section.name if s.section else ''}"
        )
        y -= 18
        if y < 80:
            c.showPage()
            y = 800

    c.save()
    return response

# -----------------------------Class Module------------------------------------
def class_list(request):
    classes = SchoolClass.objects.all()
    return render(request, 'classes/class_list.html', {'classes': classes})

#------------------------------Add/Edit/Delete Class---------------------------
def add_class(request):
    if request.method == 'POST':
        SchoolClass.objects.create(
            name=request.POST.get('name'),
            description=request.POST.get('description'),
            status=request.POST.get('status')
        )
        return redirect('class_list')
    return render(request, 'classes/add_class.html')


def edit_class(request, id):
    cls = get_object_or_404(SchoolClass, id=id)
    if request.method == 'POST':
        cls.name = request.POST.get('name')
        cls.description = request.POST.get('description')
        cls.status = request.POST.get('status')
        cls.save()
        return redirect('class_list')
    return render(request, 'classes/edit_class.html', {'cls': cls})


def delete_class(request, id):
    cls = get_object_or_404(SchoolClass, id=id)
    cls.delete()
    return redirect('class_list')

# ------------------------------Timetable List--------------------------------
def timetable_list(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()
    timetables = TimeTable.objects.select_related('school_class','section')

    selected_class = request.GET.get('class')
    selected_section = request.GET.get('section')

    if selected_class:
        timetables = timetables.filter(school_class_id=selected_class)
        sections = sections.filter(school_class_id=selected_class)

    if selected_section:
        timetables = timetables.filter(section_id=selected_section)

    return render(request, 'timetable/timetable_list.html', {
        'timetables': timetables,
        'classes': classes,
        'sections': sections,
        'selected_class': selected_class,
        'selected_section': selected_section
    })


def add_timetable(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()

    if request.method == 'POST':
        TimeTable.objects.create(
            school_class_id=request.POST.get('class'),
            section_id=request.POST.get('section'),
            day=request.POST.get('day'),
            subject=request.POST.get('subject'),
            teacher_name=request.POST.get('teacher'),
            start_time=request.POST.get('start_time'),
            end_time=request.POST.get('end_time'),
        )
        return redirect('timetable_list')

    return render(request, 'timetable/add_timetable.html', {
        'classes': classes,
        'sections': sections
    })

def edit_timetable(request, id):
    t = get_object_or_404(TimeTable, id=id)
    classes = SchoolClass.objects.all()
    sections = Section.objects.filter(school_class=t.school_class)

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    teachers = Teacher.objects.all()   

    if request.method == 'POST':
        t.school_class_id = request.POST['class']
        t.section_id = request.POST['section']
        t.day = request.POST['day']
        t.subject = request.POST['subject']
        t.teacher_id = request.POST['teacher']
        t.start_time = request.POST['start_time']
        t.end_time = request.POST['end_time']
        t.save()
        return redirect('timetable_list')

    return render(request, 'timetable/edit_timetable.html', {
        't': t,
        'classes': classes,
        'sections': sections,
        'days': days,
        'teachers': teachers,   
    })


def delete_timetable(request, id):
    TimeTable.objects.filter(id=id).delete()
    return redirect('timetable_list')

#------------------------------------Section------------------------------------
def section_list(request):
    sections = Section.objects.select_related('school_class')
    return render(request, 'sections/section_list.html', {'sections': sections})

def add_section(request):
    classes = SchoolClass.objects.all()
    if request.method == 'POST':
        Section.objects.create(
            school_class_id=request.POST.get('school_class'),
            name=request.POST.get('name')
        )
        return redirect('section_list')
    return render(request, 'sections/add_section.html', {'classes': classes})

def edit_section(request, id):
    section = get_object_or_404(Section, id=id)
    classes = SchoolClass.objects.all()

    if request.method == 'POST':
        section.school_class_id = request.POST.get('school_class')
        section.name = request.POST.get('name')
        section.save()
        return redirect('section_list')

    return render(
        request,
        'sections/edit_section.html',
        {'section': section, 'classes': classes}
    )

def delete_section(request, id):
    section = get_object_or_404(Section, id=id)
    section.delete()
    return redirect('section_list')

# ------------------------------Fees Management---------------------------------
# ================= Fee Structure =================
# def fee_structure_list(request):
#     fees = FeeStructure.objects.select_related('school_class')
#     return render(request,'fees/fee_structure_list.html',{'fees':fees})


# def fee_structure_add(request):
#     classes = SchoolClass.objects.all()
#     if request.method == 'POST':
#         FeeStructure.objects.create(
#             school_class_id=request.POST.get('school_class'),
#             total_fee=request.POST.get('total_fee'),
#             description=request.POST.get('description'),
#             status=request.POST.get('status')
#         )
#         messages.success(request,"Fee structure added")
#         return redirect('fee_structure_list')

#     return render(request,'fees/fee_structure_form.html',{'classes':classes})


# def fee_structure_edit(request,id):
#     fee = get_object_or_404(FeeStructure,id=id)
#     classes = SchoolClass.objects.all()

#     if request.method == 'POST':
#         fee.school_class_id = request.POST.get('school_class')
#         fee.total_fee = request.POST.get('total_fee')
#         fee.description = request.POST.get('description')
#         fee.status = request.POST.get('status')
#         fee.save()
#         messages.success(request,"Updated successfully")
#         return redirect('fee_structure_list')

#     return render(request,'fees/fee_structure_form.html',{'fee':fee,'classes':classes})


# def fee_structure_delete(request,id):
#     FeeStructure.objects.filter(id=id).delete()
#     messages.success(request,"Deleted")
#     return redirect('fee_structure_list')



# def fee_bill_list(request):
#     bills = FeeBill.objects.select_related('student','fee_structure')
#     return render(request,'fees/fee_bill_list.html',{'bills':bills})


# def fee_bill_add(request):
#     students = Student.objects.all()
#     structures = FeeStructure.objects.filter(status='Active')

#     if request.method == 'POST':
#         FeeBill.objects.create(
#             student_id=request.POST.get('student'),
#             fee_structure_id=request.POST.get('fee_structure'),
#             paid_amount=request.POST.get('paid_amount'),
#             payment_mode=request.POST.get('payment_mode'),
#             remark=request.POST.get('remark')
#         )
#         messages.success(request,"Fee collected")
#         return redirect('fee_bill_list')

#     return render(request,'fees/fee_bill_form.html',{
#         'students':students,
#         'structures':structures
#     })

# def pending_fees(request):
#     students = Student.objects.select_related('school_class')

#     data = []
#     for s in students:
#         structure = FeeStructure.objects.filter(
#             school_class=s.school_class,
#             status='Active'
#         ).first()

#         total_fee = structure.total_fee if structure else 0
#         paid = FeeBill.objects.filter(student=s).aggregate(
#             total=Sum('paid_amount')
#         )['total'] or 0

#         pending = total_fee - paid

#         if pending > 0:
#             data.append({
#                 'student': s,
#                 'total_fee': total_fee,
#                 'paid': paid,
#                 'pending': pending
#             })

#     return render(request,'fees/pending_fees.html',{'data':data})

# def collect_fee(request, bill_id):
#     bill = get_object_or_404(FeeBill, id=bill_id)

#     if request.method == 'POST':
#         pay_amount = Decimal(request.POST.get('pay_amount'))
#         payment_mode = request.POST.get('payment_mode')
#         payment_date = request.POST.get('payment_date')
#         remark = request.POST.get('remark')

#         if pay_amount <= 0:
#             messages.error(request, "Invalid payment amount.")
#             return redirect('collect_fee', bill_id=bill.id)

#         if pay_amount > bill.pending_amount:
#             messages.error(request, "Payment cannot exceed pending fee.")
#             return redirect('collect_fee', bill_id=bill.id)

#         # Save payment
#         FeePayment.objects.create(
#             bill=bill,
#             amount=pay_amount,
#             payment_mode=payment_mode,
#             payment_date=payment_date,
#             remark=remark
#         )

#         # Update bill
#         bill.paid_amount += pay_amount
#         bill.pending_amount = bill.total_fee - bill.paid_amount
#         bill.save()

#         messages.success(request, "Fee collected successfully.")
#         return redirect('fee_bill_list')

#     context = {
#         'bill': bill,
#         'today': timezone.now().date()
#     }
#     return render(request, 'fees/collect_fee.html', context)

# ---------- Fee Structure ----------
def fee_structure_list(request):
    fees = FeeStructure.objects.select_related('school_class')
    return render(request, 'fees/fee_structure_list.html', {'fees': fees})


def fee_structure_add(request):
    classes = SchoolClass.objects.all()

    if request.method == 'POST':
        FeeStructure.objects.create(
            school_class_id=request.POST['school_class'],
            total_fee=request.POST['total_fee'],
            description=request.POST.get('description', ''),
            status=request.POST['status']
        )
        messages.success(request, "Fee structure added")
        return redirect('fee_structure_list')

    return render(request, 'fees/fee_structure_form.html', {'classes': classes})

def fee_structure_edit(request, id):
    fee = get_object_or_404(FeeStructure, id=id)
    classes = SchoolClass.objects.all()

    if request.method == 'POST':
        fee.school_class_id = request.POST['school_class']
        fee.total_fee = request.POST['total_fee']
        fee.description = request.POST.get('description', '')
        fee.status = request.POST['status']
        fee.save()

        messages.success(request, "Fee structure updated successfully")
        return redirect('fee_structure_list')

    return render(request, 'fees/fee_structure_form.html', {
        'fee': fee,
        'classes': classes,
        'title': 'Edit Fee Structure'
    })


# def fee_structure_delete(request, id):
#     fee = get_object_or_404(FeeStructure, id=id)
#     fee.delete()
#     messages.success(request, "Fee structure deleted successfully")
#     return redirect('fee_structure_list')


# ---------- Fee Bill ----------


def fee_bill_list(request):
    bills = FeeBill.objects.select_related('student').prefetch_related('months', 'payments')
    return render(request, 'fees/fee_bill_list.html', {'bills': bills})


from django.http import JsonResponse

# def fee_bill_add(request):
#     structures = FeeStructure.objects.filter(status='Active')
#     classes = SchoolClass.objects.all()

#     # 🔥 UNIQUE SECTIONS ONLY (A, B, C, D)
#     sections = (
#         Section.objects
#         .values_list('name', flat=True)
#         .distinct()
#         .order_by('name')
#     )

#     if request.method == 'POST':
#         structure = get_object_or_404(FeeStructure, id=request.POST['fee_structure'])

#         FeeBill.objects.create(
#             student_id=request.POST['student'],
#             fee_structure=structure,
#             total_fee=structure.total_fee
#         )
#         messages.success(request, "Fee bill created successfully")
#         return redirect('fee_bill_list')

#     return render(request, 'fees/fee_bill_form.html', {
#         'structures': structures,
#         'classes': classes,
#         'sections': sections
#     })

# -------------------------------
# Fee Bill Add
# -------------------------------
def fee_bill_add(request):
    structures = FeeStructure.objects.filter(status='Active')
    classes = SchoolClass.objects.all()
    sections = Section.objects.values_list('name', flat=True).distinct()

    months = [
        "January", "February", "March",
        "April", "May", "June", "July",
        "August", "September", "October",
        "November", "December"
    ]

    if request.method == 'POST':
        student_id = request.POST.get('student')
        structure_id = request.POST.get('fee_structure')
        selected_months = request.POST.getlist('months[]')

        structure = get_object_or_404(FeeStructure, id=structure_id)

        # ✅ CREATE ONLY ONE BILL
        bill = FeeBill.objects.create(
            student_id=student_id,
            fee_structure=structure,
            total_fee=structure.total_fee * len(selected_months)
        )

        # ✅ SAVE MONTHS
        for m in selected_months:
            FeeBillMonth.objects.create(
                bill=bill,
                month=m.strip()
            )

        messages.success(request, "Fee bill generated successfully")
        return redirect('fee_bill_list')

    return render(request, 'fees/fee_bill_form.html', {
        'structures': structures,
        'classes': classes,
        'sections': sections,
        'months': months
    })



def get_paid_months(request, student_id):
    paid_months = []

    bills = FeeBill.objects.filter(student_id=student_id)

    for bill in bills:
        total_paid = bill.payments.aggregate(
            total=Sum('amount')
        )['total'] or 0

        # ✅ FULL BILL PAID
        if total_paid >= bill.total_fee:
            months = bill.months.values_list('month', flat=True)
            paid_months.extend([m.strip() for m in months])

    return JsonResponse({
        'paid_months': list(set(paid_months))  # ✅ no duplicates
    })


# 🔥 AJAX VIEW (LIVE FILTER)
def get_filtered_students(request):
    name = request.GET.get('name', '').strip()
    class_id = request.GET.get('class_id', '')
    section_name = request.GET.get('section_id', '')

    students = Student.objects.select_related('school_class', 'section')

    if name:
        students = students.filter(name__icontains=name)

    if class_id:
        students = students.filter(school_class__id=class_id)

    if section_name:
        students = students.filter(section__name=section_name)

    data = []
    for s in students:
        data.append({
            "id": s.id,
            "text": f"{s.name} ({s.school_class.name} - {s.section.name})"
        })

    return JsonResponse(data, safe=False)


from xhtml2pdf import pisa
from django.template.loader import get_template
def fee_receipt_download(request, bill_id):
    bill = get_object_or_404(FeeBill, id=bill_id)
    payment = bill.payments.last()

    months = bill.months.all()

    # ✅ Convert amounts to float
    total_fee = float(bill.total_fee)
    paid_amount = float(payment.amount if payment else 0)
    pending_amount = total_fee - paid_amount

    context = {
        'bill': bill,
        'payment': payment,
        'months': months,
        'total_fee': total_fee,
        'paid_amount': paid_amount,
        'pending_amount': pending_amount,
    }

    # Generate PDF using template (xhtml2pdf)
    template_path = 'fees/fee_receipt.html'
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="fee_receipt_{bill.id}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('PDF generation failed')
    return response



# ---------- Collect Fee ----------
def collect_fee(request, bill_id):
    bill = get_object_or_404(FeeBill, id=bill_id)

    if request.method == 'POST':
        amount = Decimal(request.POST['pay_amount'])

        if amount <= 0 or amount > bill.pending_amount():
            messages.error(request, "Invalid payment amount")
            return redirect('collect_fee', bill_id=bill.id)

        FeePayment.objects.create(
            bill=bill,
            amount=amount,
            payment_mode=request.POST['payment_mode'],
            payment_date=request.POST['payment_date'],
            remark=request.POST.get('remark', '')
        )

        messages.success(request, "Fee collected successfully")
        return redirect('fee_bill_list')

    return render(request, 'fees/collect_fee.html', {
        'bill': bill,
        'today': timezone.now().date()
    })


# ---------- Pending Fees ----------
def pending_fees(request):
    bills = FeeBill.objects.all()
    pending_bills = [b for b in bills if b.pending_amount() > 0]
    return render(request, 'fees/pending_fees.html', {'bills': pending_bills})



#Exam Management
def exam_list(request):
    exams = Exam.objects.all()
    return render(request, 'exams/exam_list.html', {'exams': exams})


def add_exam(request):
    classes = SchoolClass.objects.all()
    sections = Section.objects.all()

    if request.method == 'POST':
        Exam.objects.create(
            name=request.POST['name'],
            school_class_id=request.POST['school_class'],
            section_id=request.POST['section'],
            exam_date=request.POST['exam_date']
        )
        return redirect('exam_list')

    return render(request, 'exams/exam_form.html', {
        'classes': classes,
        'sections': sections
    })


def upload_result(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    students = Student.objects.filter(
        school_class=exam.school_class,
        section=exam.section
    )

    if request.method == 'POST':
        for s in students:
            marks = request.POST.get(f"marks_{s.id}")
            if marks:
                ExamResult.objects.create(
                    exam=exam,
                    student=s,
                    subject=request.POST['subject'],
                    marks=marks
                )
        return redirect('exam_list')

    return render(request, 'exams/upload_result.html', {
        'exam': exam,
        'students': students
    })

def upload_result(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id)
    students = Student.objects.filter(
        school_class=exam.school_class,
        section=exam.section
    )

    if request.method == 'POST':
        for s in students:
            marks = request.POST.get(f"marks_{s.id}")
            if marks:
                ExamResult.objects.create(
                    exam=exam,
                    student=s,
                    subject=request.POST['subject'],
                    marks=marks
                )
        return redirect('exam_list')

    return render(request, 'exams/upload_result.html', {
        'exam': exam,
        'students': students
    })

def report_card(request, student_id, exam_id):
    student = get_object_or_404(Student, id=student_id)
    exam = get_object_or_404(Exam, id=exam_id)

    results = ExamResult.objects.filter(student=student, exam=exam)

    total = results.aggregate(
        total_marks=Sum('marks'),
        max_marks=Sum('max_marks')
    )

    percentage = 0
    if total['max_marks']:
        percentage = (total['total_marks'] / total['max_marks']) * 100

    grade = 'Fail'
    if percentage >= 90: grade = 'A+'
    elif percentage >= 75: grade = 'A'
    elif percentage >= 60: grade = 'B'
    elif percentage >= 40: grade = 'C'

    return render(request, 'exams/report_card.html', {
        'student': student,
        'exam': exam,
        'results': results,
        'percentage': round(percentage, 2),
        'grade': grade
    })

# ================= TEACHER =================


def homework_list(request):
    homework = Homework.objects.all().order_by('-created_at')
    return render(request, 'homework/homework_list.html', {'homework': homework})


def add_homework(request):
    if request.method == 'POST':
        Homework.objects.create(
            title=request.POST['title'],
            description=request.POST['description'],
            subject=request.POST['subject'],
            homework_file=request.FILES.get('homework_file'),
            due_date=request.POST['due_date']
        )
        return redirect('homework_list')

    return render(request, 'homework/add_homework.html')


def edit_homework(request, pk):
    hw = get_object_or_404(Homework, pk=pk)

    if request.method == 'POST':
        hw.title = request.POST['title']
        hw.description = request.POST['description']
        hw.subject = request.POST['subject']
        hw.due_date = request.POST['due_date']

        if request.FILES.get('homework_file'):
            hw.homework_file = request.FILES.get('homework_file')

        hw.save()
        return redirect('homework_list')

    return render(request, 'homework/edit_homework.html', {'hw': hw})


def delete_homework(request, pk):
    hw = get_object_or_404(Homework, pk=pk)
    hw.delete()
    return redirect('homework_list')

# -------- NOTICE LIST --------
# List
def notice_list(request):
    notices = Notice.objects.all().order_by('-id')
    return render(request, 'notice/notice_list.html', {'notices': notices})


# Add
def add_notice(request):
    if request.method == 'POST':
        Notice.objects.create(
            title=request.POST['title'],
            description=request.POST['description'],
            notice_date=request.POST['notice_date']
        )
        return redirect('notice_list')

    return render(request, 'notice/add_notice.html')


# Edit
def edit_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)

    if request.method == 'POST':
        notice.title = request.POST['title']
        notice.description = request.POST['description']
        notice.notice_date = request.POST['notice_date'] 
        notice.save()
        return redirect('notice_list')

    return render(request, 'notice/edit_notice.html', {'notice': notice})


# Delete
def delete_notice(request, pk):
    notice = get_object_or_404(Notice, pk=pk)
    notice.delete()
    return redirect('notice_list')


# ROUTES
def route_list(request):
    return render(request, 'transport/route_list.html', {
        'routes': Route.objects.all()
    })

def route_add(request):
    if request.method == 'POST':
        Route.objects.create(
            name=request.POST['name'],
            start_point=request.POST['start'],
            end_point=request.POST['end'],
            monthly_fee=request.POST['fee']
        )
        return redirect('route_list')
    return render(request, 'transport/route_form.html')


# DRIVERS
def driver_list(request):
    return render(request, 'transport/driver_list.html', {
        'drivers': Driver.objects.all()
    })

def driver_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        mobile = request.POST.get('mobile')
        license_no = request.POST.get('license_no')   # ✅ FIXED

        Driver.objects.create(
            name=name,
            mobile=mobile,
            license_no=license_no
        )
        return redirect('driver_list')

    return render(request, 'transport/driver_form.html')


# VEHICLES
def vehicle_list(request):
    return render(request, 'transport/vehicle_list.html', {
        'vehicles': Vehicle.objects.all(),
        'drivers': Driver.objects.all()
    })

# def vehicle_add(request):
#     if request.method == 'POST':
#         Vehicle.objects.create(
#             vehicle_no=request.POST['vehicle_no'],
#             capacity=request.POST['capacity'],
#             driver_id=request.POST['driver']
#         )
#         return redirect('vehicle_list')
#     return render(request, 'transport/vehicle_form.html')

def vehicle_add(request):
    drivers = Driver.objects.all()   # 👈 IMPORTANT

    if request.method == 'POST':
        vehicle_no = request.POST.get('vehicle_no')
        capacity = request.POST.get('capacity')
        driver_id = request.POST.get('driver')

        driver = Driver.objects.get(id=driver_id) if driver_id else None

        Vehicle.objects.create(
            vehicle_no=vehicle_no,
            capacity=capacity,
            driver=driver
        )

        return redirect('vehicle_list')

    context = {
        'drivers': drivers
    }
    return render(request, 'transport/vehicle_form.html', context)


# STUDENT ALLOCATION
def allocation_list(request):
    return render(request, 'transport/allocation_list.html', {
        'allocations': TransportAllocation.objects.select_related('student','route')
    })

def allocation_add(request):
    if request.method == 'POST':
        TransportAllocation.objects.create(
            student_id=request.POST['student'],
            route_id=request.POST['route'],
            vehicle_id=request.POST['vehicle'],
            # start_date=request.POST['start_date']
        )
        messages.success(request, "Transport Allocated")
        return redirect('allocation_list')

    return render(request, 'transport/allocation_form.html', {
        'students': Student.objects.all(),
        'routes': Route.objects.all(),
        'vehicles': Vehicle.objects.all()
    })

# edit/delete of Transport Management
def route_edit(request, id):
    route = get_object_or_404(Route, id=id)

    if request.method == 'POST':
        route.name = request.POST.get('name')
        route.start_point = request.POST.get('start_point')
        route.end_point = request.POST.get('end_point')
        route.save()
        messages.success(request, "Route updated successfully")
        return redirect('route_list')

    return render(request, 'transport/route_form.html', {'route': route})


def route_delete(request, id):
    route = get_object_or_404(Route, id=id)
    route.delete()
    messages.success(request, "Route deleted")
    return redirect('route_list')


# ================= DRIVER =================
def driver_edit(request, id):
    driver = get_object_or_404(Driver, id=id)

    if request.method == 'POST':
        driver.name = request.POST.get('name')
        driver.mobile = request.POST.get('mobile')
        driver.license_no = request.POST.get('license_no')
        driver.save()
        messages.success(request, "Driver updated")
        return redirect('driver_list')

    return render(request, 'transport/driver_form.html', {'driver': driver})


def driver_delete(request, id):
    driver = get_object_or_404(Driver, id=id)
    driver.delete()
    messages.success(request, "Driver deleted")
    return redirect('driver_list')


# ================= VEHICLE =================
def vehicle_edit(request, id):
    vehicle = get_object_or_404(Vehicle, id=id)
    drivers = Driver.objects.all()

    if request.method == 'POST':
        vehicle.vehicle_no = request.POST.get('vehicle_no')
        vehicle.capacity = request.POST.get('capacity')
        vehicle.driver_id = request.POST.get('driver')
        vehicle.save()
        messages.success(request, "Vehicle updated")
        return redirect('vehicle_list')

    return render(request, 'transport/vehicle_form.html', {
        'vehicle': vehicle,
        'drivers': drivers
    })


def vehicle_delete(request, id):
    vehicle = get_object_or_404(Vehicle, id=id)
    vehicle.delete()
    messages.success(request, "Vehicle deleted")
    return redirect('vehicle_list')


# ================= ALLOCATION =================
def allocation_edit(request, id):
    allocation = get_object_or_404(TransportAllocation, id=id)
    students = Student.objects.all()
    routes = Route.objects.all()
    vehicles = Vehicle.objects.all()

    if request.method == 'POST':
        allocation.student_id = request.POST.get('student')
        allocation.route_id = request.POST.get('route')
        allocation.vehicle_id = request.POST.get('vehicle')
        allocation.save()
        messages.success(request, "Allocation updated")
        return redirect('allocation_list')

    return render(request, 'transport/allocation_form.html', {
        'allocation': allocation,
        'students': students,
        'routes': routes,
        'vehicles': vehicles
    })


def allocation_delete(request, id):
    allocation = get_object_or_404(TransportAllocation, id=id)
    allocation.delete()
    messages.success(request, "Allocation deleted")
    return redirect('allocation_list')

#Library Management
def book_list(request):
    books = Book.objects.all()
    return render(request, 'library/book_list.html', {'books': books})


def book_add(request):
    if request.method == 'POST':
        Book.objects.create(
            title=request.POST['title'],
            author=request.POST['author'],
            isbn=request.POST['isbn'],
            quantity=request.POST['quantity']
        )
        messages.success(request, "Book added")
        return redirect('book_list')

    return render(request, 'library/book_form.html')



def book_edit(request, id):
    book = get_object_or_404(Book, id=id)

    if request.method == 'POST':
        book.title = request.POST['title']
        book.author = request.POST['author']
        book.isbn = request.POST['isbn']
        book.quantity = request.POST['quantity']
        book.save()
        messages.success(request, "Book updated")
        return redirect('book_list')

    return render(request, 'library/book_form.html', {'book': book})


def book_delete(request, id):
    Book.objects.filter(id=id).delete()
    messages.success(request, "Book deleted")
    return redirect('book_list')



def issue_list(request):
    issues = BookIssue.objects.all()
    return render(request, 'library/issue_list.html', {'issues': issues})


def issue_book(request):
    books = Book.objects.all()
    students = Student.objects.all()
    teachers = Teacher.objects.all()

    if request.method == 'POST':
        book = Book.objects.get(id=request.POST['book'])

        if book.quantity <= 0:
            messages.error(request, "Book not available")
            return redirect('issue_book')

        due_date = request.POST.get('due_date')

        if not due_date:
            messages.error(request, "Due date is required")
            return redirect('issue_book')

        issue = BookIssue.objects.create(
            book=book,
            student_id=request.POST.get('student') or None,
            teacher_id=request.POST.get('teacher') or None,
            due_date=due_date
        )

        book.quantity -= 1
        book.save()

        messages.success(request, "Book issued successfully")
        return redirect('issue_list')

    return render(request, 'library/issue_form.html', {
        'books': books,
        'students': students,
        'teachers': teachers
    })


# def return_book(request, id):
#     issue = get_object_or_404(BookIssue, id=id)

#     if not issue.is_returned:
#         issue.is_returned = True
#         issue.return_date = date.today()
#         issue.save()

#         issue.book.quantity += 1
#         issue.book.save()

#     messages.success(request, "Book returned")
#     return redirect('issue_list')

def return_book(request, id):
    issue = get_object_or_404(BookIssue, id=id)

    if not issue.is_returned:
        issue.is_returned = True
        issue.return_date = date.today()

        # Fine calculation
        if issue.due_date and issue.return_date > issue.due_date:
            days_late = (issue.return_date - issue.due_date).days
            issue.fine_amount = days_late * 10

        issue.save()

        issue.book.quantity += 1
        issue.book.save()

    messages.success(request, "Book returned successfully")
    return redirect('issue_list')


#Warden Management
def warden_list(request):
    wardens = Warden.objects.all().order_by('-id')
    return render(request, 'warden/warden_list.html', {'wardens': wardens})


def warden_add(request):
    if request.method == "POST":
        Warden.objects.create(
            name=request.POST['name'],
            mobile=request.POST['mobile'],
            email=request.POST.get('email'),
            address=request.POST.get('address'),
            join_date=request.POST['join_date'],
            is_active=True if request.POST.get('is_active') else False
        )
        return redirect('warden_list')

    return render(request, 'warden/warden_form.html')


def warden_edit(request, id):
    warden = get_object_or_404(Warden, id=id)

    if request.method == "POST":
        warden.name = request.POST['name']
        warden.mobile = request.POST['mobile']
        warden.email = request.POST.get('email')
        warden.address = request.POST.get('address')
        warden.join_date = request.POST['join_date']
        warden.is_active = True if request.POST.get('is_active') else False
        warden.save()
        return redirect('warden_list')

    return render(request, 'warden/warden_form.html', {'warden': warden})


def warden_delete(request, id):
    warden = get_object_or_404(Warden, id=id)
    warden.delete()
    return redirect('warden_list')

# ================= HOSTEL LIST =================
def hostel_list(request):
    return render(request, 'hostel/hostel_list.html', {
        'hostels': Hostel.objects.select_related('warden')
    })


# ================= ADD / EDIT HOSTEL =================
def hostel_form(request, id=None):
    hostel = Hostel.objects.filter(id=id).first()
    wardens = Warden.objects.all()

    if request.method == 'POST':
        Hostel.objects.update_or_create(
            id=id,
            defaults={
                'name': request.POST.get('name'),
                'hostel_type': request.POST.get('hostel_type'),  # ✅ FIXED
                'address': request.POST.get('address'),
                'warden_id': request.POST.get('warden'),
                'total_rooms': request.POST.get('total_rooms'),
            }
        )
        return redirect('hostel_list')

    return render(request, 'hostel/hostel_form.html', {
        'hostel': hostel,
        'wardens': wardens,
        'hostel_types': Hostel.HOSTEL_TYPE_CHOICES  # ✅ send choices to HTML
    })


# ================= DELETE HOSTEL =================
def hostel_delete(request, id):
    Hostel.objects.filter(id=id).delete()
    return redirect('hostel_list')


# ================= ROOMS =================
def room_list(request):
    return render(request, 'hostel/room_list.html', {
        'rooms': Room.objects.all()
    })

def room_form(request, id=None):
    room = Room.objects.filter(id=id).first()
    hostels = Hostel.objects.all()
    error = None

    if request.method == 'POST':
        hostel_id = request.POST.get('hostel')
        room_no = request.POST.get('room_no')
        capacity = int(request.POST.get('capacity', 0))
        occupied = int(request.POST.get('occupied', 0))

        # 🔴 VALIDATION
        if occupied > capacity:
            error = "Occupied beds cannot be greater than capacity."
        else:
            Room.objects.update_or_create(
                id=id,
                defaults={
                    'hostel_id': hostel_id,
                    'room_no': room_no,
                    'capacity': capacity,
                    'occupied': occupied,
                }
            )
            return redirect('room_list')

    return render(request, 'hostel/room_form.html', {
        'room': room,
        'hostels': hostels,
        'error': error
    })


def room_delete(request, id):
    Room.objects.filter(id=id).delete()
    return redirect('room_list')


# ================= ALLOCATION LIST =================
# def room_allocation_list(request):
#     allocations = RoomAllocation.objects.select_related(
#         'student', 'room', 'room__hostel'
#     )
#     return render(request, 'hostel/room_allocation_list.html', {
#         'allocations': allocations
#     })




# ================= ALLOCATION ADD =================
# def room_allocation_add(request):
#     students = Student.objects.all()
#     hostels = Hostel.objects.all()
#     rooms = Room.objects.all()

#     if request.method == 'POST':
#         student_id = request.POST.get('student')
#         hostel_id = request.POST.get('hostel')
#         room_id = request.POST.get('room')

#         # One student → one bed rule
#         if RoomAllocation.objects.filter(student_id=student_id).exists():
#             messages.error(request, "This student is already allocated.")
#             return redirect('room_allocation_add')

#         RoomAllocation.objects.create(
#             student_id=student_id,
#             hostel_id=hostel_id,
#             room_id=room_id
#         )

#         messages.success(request, "Student allocated successfully.")
#         return redirect('room_allocation_list')

#     return render(request, 'hostel/room_allocation_add.html', {
#         'students': students,
#         'hostels': hostels,
#         'rooms': rooms
#     })



# ================= ALLOCATION DELETE =================
# def room_allocation_delete(request, id):
#     RoomAllocation.objects.filter(id=id).delete()
#     messages.success(request, "Allocation removed successfully.")
#     return redirect('room_allocation_list')

from django.db.models import Subquery


def room_allocation_list(request, id=None):
    allocations = RoomAllocation.objects.select_related('student', 'room', 'hostel')

    # 🔥 EXCLUDE TRANSPORT STUDENTS
    transport_students = TransportAllocation.objects.filter(
        active=True
    ).values('student_id')

    students = Student.objects.exclude(id__in=transport_students)

    hostels = Hostel.objects.all()
    rooms = Room.objects.all()

    allocation = None
    is_edit = False

    if id:
        allocation = get_object_or_404(RoomAllocation, id=id)
        is_edit = True

    if request.method == 'POST':
        student_id = request.POST.get('student')
        hostel_id = request.POST.get('hostel')
        room_id = request.POST.get('room')

        # ❗ SAFETY CHECK (VERY IMPORTANT)
        if TransportAllocation.objects.filter(
            student_id=student_id, active=True
        ).exists():
            messages.error(
                request,
                "This student has transport facility and cannot be allocated hostel."
            )
            return redirect('room_allocation_list')

        # ❗ DUPLICATE ALLOCATION CHECK
        qs = RoomAllocation.objects.filter(student_id=student_id)
        if is_edit:
            qs = qs.exclude(id=id)

        if qs.exists():
            messages.error(request, "This student is already allocated.")
            return redirect('room_allocation_list')

        if is_edit:
            allocation.student_id = student_id
            allocation.hostel_id = hostel_id
            allocation.room_id = room_id
            allocation.save()
            messages.success(request, "Allocation updated successfully.")
        else:
            RoomAllocation.objects.create(
                student_id=student_id,
                hostel_id=hostel_id,
                room_id=room_id
            )
            messages.success(request, "Student allocated successfully.")

        return redirect('room_allocation_list')

    return render(request, 'hostel/room_allocation.html', {
        'allocations': allocations,
        'students': students,
        'hostels': hostels,
        'rooms': rooms,
        'allocation': allocation,
        'is_edit': is_edit
    })



def room_allocation_add(request):
    return room_allocation_list(request)


def room_allocation_delete(request, id):
    RoomAllocation.objects.filter(id=id).delete()
    messages.success(request, "Allocation removed successfully.")
    return redirect('room_allocation_list')



# ================= FEES =================
def fee_list(request):
    return render(request, 'hostel/fee_list.html', {
        'fees': HostelFee.objects.all()
    })

def fee_form(request):

    # 🔥 ONLY HOSTEL ALLOCATED STUDENTS
    allocated_students = RoomAllocation.objects.values('student_id')

    students = Student.objects.filter(
        id__in=Subquery(allocated_students)
    )

    hostels = Hostel.objects.all()

    if request.method == 'POST':
        HostelFee.objects.create(
            student_id=request.POST['student'],
            # hostel_id=request.POST['hostel'],
            amount=request.POST['amount'],
            paid_date=request.POST['paid_date'],
        )
        return redirect('hostel_fee_list')

    return render(request, 'hostel/fee_form.html', {
        'students': students,
        'hostels': hostels
    })

# ================= DELETE HOSTEL FEE =================
def fee_delete(request, id):
    fee = get_object_or_404(HostelFee, id=id)
    fee.delete()
    return redirect('hostel_fee_list')